# Excel loading with multi-sheet support and bronze persistence

import os
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from src.core.config import BRONZE_OUTPUT_DIR
from src.core.exceptions import FileReadError


EXCEL_MAX_ROWS_PER_SHEET = 200_000
EXCEL_EMPTY_ROW_BREAK_THRESHOLD = 5_000


def _read_xlsx_sheet_streaming(file_path: str, sheet_name: str) -> pd.DataFrame:
    """Stream-read one xlsx sheet using openpyxl to avoid pandas/openpyxl hangs on large sheets."""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]

        header = None
        rows = []
        empty_streak = 0

        for row in ws.iter_rows(values_only=True):
            values = [v for v in row]
            is_empty = all(v is None for v in values)

            if header is None:
                if is_empty:
                    continue
                header = [str(v).strip() if v is not None else "" for v in values]
                continue

            if is_empty:
                empty_streak += 1
                if empty_streak >= EXCEL_EMPTY_ROW_BREAK_THRESHOLD:
                    print(
                        f"[INGEST_TRACE] load_excel() sheet={sheet_name} "
                        f"breaking after {empty_streak} consecutive empty rows",
                        flush=True,
                    )
                    break
                continue

            empty_streak = 0
            rows.append(values)
            if len(rows) >= EXCEL_MAX_ROWS_PER_SHEET:
                print(
                    f"[INGEST_TRACE] load_excel() sheet={sheet_name} "
                    f"truncated at {EXCEL_MAX_ROWS_PER_SHEET} rows",
                    flush=True,
                )
                break

        if header is None:
            return pd.DataFrame()

        width = len(header)
        normalized_rows = []
        for r in rows:
            if len(r) < width:
                normalized_rows.append(list(r) + [None] * (width - len(r)))
            else:
                normalized_rows.append(list(r[:width]))

        # Ensure non-empty and unique column names.
        seen: dict[str, int] = {}
        safe_cols = []
        for idx, col in enumerate(header):
            base = col if col else f"column_{idx + 1}"
            if base in seen:
                seen[base] += 1
                safe_cols.append(f"{base}_{seen[base]}")
            else:
                seen[base] = 1
                safe_cols.append(base)

        return pd.DataFrame(normalized_rows, columns=safe_cols)
    finally:
        wb.close()


def load_excel(file_path: str, file_ext: Optional[str] = None) -> dict[str, pd.DataFrame]:
    """Loads all sheets from an Excel file into a dict of DataFrames.

    `file_ext` should be the original uploaded extension (e.g. ".xlsx"), because
    temp file paths may not preserve the original suffix.
    """
    ext = (file_ext or os.path.splitext(file_path)[1]).lower()
    if ext in {".xlsx", ".xlsm"}:
        engines = ["openpyxl"]
    elif ext == ".xls":
        engines = ["xlrd"]
    else:
        engines = [None, "openpyxl", "xlrd"]

    failures: list[str] = []
    print(
        f"[INGEST_TRACE] load_excel() file_path={file_path} file_ext={file_ext} "
        f"engine_candidates={engines}",
        flush=True,
    )
    for engine in engines:
        xls = None
        try:
            print(f"[INGEST_TRACE] load_excel() trying engine={engine or 'auto'}", flush=True)
            xls = pd.ExcelFile(file_path, engine=engine)
            print(
                f"[INGEST_TRACE] load_excel() engine={engine or 'auto'} opened workbook "
                f"sheets={len(xls.sheet_names)}",
                flush=True,
            )
            frames: dict[str, pd.DataFrame] = {}
            for sheet in xls.sheet_names:
                print(
                    f"[INGEST_TRACE] load_excel() reading sheet={sheet} engine={engine or 'auto'}",
                    flush=True,
                )
                if engine == "openpyxl":
                    df = _read_xlsx_sheet_streaming(file_path, sheet)
                else:
                    df = pd.read_excel(xls, sheet_name=sheet, engine=engine)
                print(
                    f"[INGEST_TRACE] load_excel() sheet={sheet} loaded rows={len(df)} cols={len(df.columns)}",
                    flush=True,
                )
                frames[sheet] = df
            return frames
        except Exception as e:
            engine_label = engine or "auto"
            print(
                f"[INGEST_TRACE] load_excel() engine={engine_label} failed error={e}",
                flush=True,
            )
            failures.append(f"{engine_label}: {e}")
        finally:
            if xls is not None:
                try:
                    xls.close()
                except Exception:
                    pass

    raise FileReadError(
        file_path,
        (
            "Failed to parse workbook with available Excel engines. "
            "Install Excel dependencies: openpyxl (for .xlsx/.xlsm) and xlrd (for .xls). "
            f"Attempted [{', '.join(failures)}]"
        ),
    )


def save_excel_sheet_to_bronze(df: pd.DataFrame, original_file: str, sheet_name: str) -> dict:
    """Saves an Excel sheet to the bronze layer as a CSV file."""
    os.makedirs(BRONZE_OUTPUT_DIR, exist_ok=True)

    file_stem = os.path.splitext(os.path.basename(original_file))[0]
    safe_sheet = str(sheet_name).strip().lower().replace(" ", "_").replace("-", "_")
    output_filename = f"{file_stem}__{safe_sheet}_bronze.csv"
    output_path = os.path.join(BRONZE_OUTPUT_DIR, output_filename)

    df.to_csv(output_path, index=False)

    return {
        "rows": len(df),
        "columns": len(df.columns),
        "output_path": output_path,
        "storage_format": "csv",
        "sheet_name": sheet_name,
        "timestamp": datetime.now().isoformat(),
    }
